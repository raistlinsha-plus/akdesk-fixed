from __future__ import annotations

import csv
import io
import re
from datetime import date, datetime
from pathlib import Path
from typing import Any

from pydantic import ValidationError

from .credit import evaluate_credit_observation
from .models import CreditObservationCreate, CreditObservationItem


CREDIT_IMPORT_FIELDS = [
    "bond_code",
    "bond_name",
    "issuer",
    "canonical_issuer",
    "sector",
    "region",
    "issuer_type",
    "maturity_date",
    "put_date",
    "rating",
    "rating_date",
    "yield_pct",
    "yield_observation_date",
    "benchmark_tenor",
    "benchmark_yield_pct",
    "benchmark_observation_date",
    "source_note",
    "next_review_date",
    "watch_status",
]

FIELD_LABELS = {
    "bond_code": "债券代码",
    "bond_name": "债券名称",
    "issuer": "发行人原名",
    "canonical_issuer": "主体归一名",
    "sector": "行业",
    "region": "地区",
    "issuer_type": "主体类型",
    "maturity_date": "到期日",
    "put_date": "回售日",
    "rating": "债项评级",
    "rating_date": "评级日期",
    "yield_pct": "到期收益率(%)",
    "yield_observation_date": "收益率日期",
    "benchmark_tenor": "基准期限",
    "benchmark_yield_pct": "基准收益率(%)",
    "benchmark_observation_date": "基准日期",
    "source_note": "来源说明",
    "next_review_date": "下次复盘日",
    "watch_status": "观察状态",
}

ALIASES = {
    "bond_code": ["债券代码", "证券代码", "代码", "bondcode", "code"],
    "bond_name": ["债券名称", "证券名称", "简称", "bondname", "name"],
    "issuer": ["发行人原名", "发行人", "主体名称", "issuer"],
    "canonical_issuer": ["主体归一名", "归一主体", "canonicalissuer"],
    "sector": ["行业", "sector", "industry"],
    "region": ["地区", "区域", "region"],
    "issuer_type": ["主体类型", "企业性质", "issuertype"],
    "maturity_date": ["到期日", "兑付日", "maturitydate"],
    "put_date": ["回售日", "putdate"],
    "rating": ["债项评级", "评级", "rating"],
    "rating_date": ["评级日期", "ratingdate"],
    "yield_pct": ["到期收益率(%)", "到期收益率", "收益率", "yieldpct", "yield"],
    "yield_observation_date": ["收益率日期", "行情日期", "观察日期", "yieldobservationdate"],
    "benchmark_tenor": ["基准期限", "benchmarktenor"],
    "benchmark_yield_pct": ["基准收益率(%)", "基准收益率", "benchmarkyieldpct"],
    "benchmark_observation_date": ["基准日期", "benchmarkobservationdate"],
    "source_note": ["来源说明", "数据来源", "source", "sourcenote"],
    "next_review_date": ["下次复盘日", "复盘日", "nextreviewdate"],
    "watch_status": ["观察状态", "状态", "watchstatus"],
}

DATE_FIELDS = {
    "maturity_date",
    "put_date",
    "rating_date",
    "yield_observation_date",
    "benchmark_observation_date",
    "next_review_date",
}
NUMBER_FIELDS = {"yield_pct", "benchmark_yield_pct"}
STATUS_ALIASES = {
    "正常观察": "active",
    "持续观察": "active",
    "关注": "watch",
    "重点关注": "watch",
    "待复盘": "review",
    "复盘": "review",
    "已归档": "archived",
    "归档": "archived",
}


def _normalize_header(value: object) -> str:
    return re.sub(r"[\s_\-/（）()%]+", "", str(value or "").strip().casefold())


def suggest_mapping(headers: list[str]) -> dict[str, str]:
    normalized = {_normalize_header(header): header for header in headers}
    result: dict[str, str] = {}
    for field in CREDIT_IMPORT_FIELDS:
        for alias in [FIELD_LABELS[field], field, *ALIASES.get(field, [])]:
            header = normalized.get(_normalize_header(alias))
            if header:
                result[field] = header
                break
    return result


def _csv_rows(content: bytes) -> tuple[list[str], list[list[Any]]]:
    text = None
    for encoding in ("utf-8-sig", "gb18030"):
        try:
            text = content.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        raise ValueError("CSV 编码无法识别，请另存为 UTF-8 或 GB18030")
    sample = text[:4096]
    try:
        dialect = csv.Sniffer().sniff(sample, delimiters=",\t;")
    except csv.Error:
        dialect = csv.excel
    rows = list(csv.reader(io.StringIO(text), dialect))
    if not rows:
        raise ValueError("文件没有表头")
    return [str(value).strip() for value in rows[0]], rows[1:]


def _xlsx_rows(content: bytes) -> tuple[list[str], list[list[Any]]]:
    try:
        from openpyxl import load_workbook
    except ImportError as exc:  # pragma: no cover - dependency gate
        raise ValueError("当前环境未安装 Excel 读取组件 openpyxl") from exc
    workbook = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
    try:
        sheet = workbook.active
        values = list(sheet.iter_rows(values_only=True, max_row=502))
    finally:
        workbook.close()
    if not values:
        raise ValueError("工作表没有表头")
    return [str(value or "").strip() for value in values[0]], [list(row) for row in values[1:]]


def _format_date(value: object) -> str | None:
    if value is None or str(value).strip() == "":
        return None
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    text = str(value).strip()
    for pattern in ("%Y-%m-%d", "%Y/%m/%d", "%Y.%m.%d", "%Y%m%d"):
        try:
            return datetime.strptime(text, pattern).date().isoformat()
        except ValueError:
            continue
    return text


def _format_number(value: object) -> float | None:
    if value is None or str(value).strip() == "":
        return None
    text = str(value).strip().replace(",", "").removesuffix("%")
    return float(text)


def _format_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def parse_credit_file(
    content: bytes,
    filename: str,
    mapping_override: dict[str, str] | None = None,
) -> dict[str, Any]:
    extension = Path(filename).suffix.lower()
    if extension == ".csv":
        headers, rows = _csv_rows(content)
    elif extension in {".xlsx", ".xlsm"}:
        headers, rows = _xlsx_rows(content)
    else:
        raise ValueError("仅支持 .csv、.xlsx 或 .xlsm 文件")
    if len(rows) > 500:
        raise ValueError("单次最多导入 500 行")
    if not any(headers):
        raise ValueError("文件表头为空")

    mapping = suggest_mapping(headers)
    if mapping_override is not None:
        mapping = {
            field: header
            for field, header in mapping_override.items()
            if field in CREDIT_IMPORT_FIELDS and header in headers
        }
    header_index = {header: index for index, header in enumerate(headers)}
    missing_required = [field for field in ("bond_code", "bond_name") if field not in mapping]
    errors: list[dict[str, Any]] = []
    valid_items: list[CreditObservationCreate] = []
    preview_rows: list[dict[str, Any]] = []

    if missing_required:
        errors.append(
            {
                "row": 1,
                "message": "缺少必填字段映射："
                + "、".join(FIELD_LABELS[field] for field in missing_required),
            }
        )

    for row_number, row in enumerate(rows, start=2):
        if not any(value is not None and str(value).strip() for value in row):
            continue
        payload: dict[str, Any] = {}
        try:
            for field, header in mapping.items():
                index = header_index[header]
                value = row[index] if index < len(row) else None
                if field in DATE_FIELDS:
                    payload[field] = _format_date(value)
                elif field in NUMBER_FIELDS:
                    payload[field] = _format_number(value)
                elif field == "benchmark_tenor":
                    payload[field] = _format_text(value) or None
                else:
                    payload[field] = _format_text(value)
            if payload.get("rating"):
                payload["rating"] = str(payload["rating"]).upper()
            status = str(payload.get("watch_status") or "active")
            payload["watch_status"] = STATUS_ALIASES.get(status, status.casefold())
            item = CreditObservationCreate(**payload)
            evaluated = evaluate_credit_observation(item)
            valid_items.append(item)
            preview_rows.append(
                {
                    "row": row_number,
                    "bond_code": item.bond_code,
                    "bond_name": item.bond_name,
                    "issuer": item.canonical_issuer or item.issuer,
                    "yield_observation_date": item.yield_observation_date,
                    "eligible": evaluated["eligible"],
                    "quality_notes": [
                        *[f"缺少 {FIELD_LABELS.get(field, field)}" for field in evaluated["missing_fields"]],
                        *evaluated["gate_reasons"],
                    ],
                }
            )
        except (ValueError, ValidationError) as exc:
            errors.append({"row": row_number, "message": str(exc)})

    return {
        "filename": filename,
        "headers": headers,
        "mapping": mapping,
        "fields": [
            {
                "field": field,
                "label": FIELD_LABELS[field],
                "required": field in {"bond_code", "bond_name"},
            }
            for field in CREDIT_IMPORT_FIELDS
        ],
        "total_rows": len(valid_items) + len([error for error in errors if error["row"] > 1]),
        "valid_rows": len(valid_items),
        "invalid_rows": len([error for error in errors if error["row"] > 1]),
        "errors": errors[:50],
        "preview_rows": preview_rows[:20],
        "items": valid_items,
        "notice": "导入只保存用户提供的字段；未提供来源或日期时会保留为空，并阻断机械利差计算。",
    }


def credit_import_template() -> str:
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([FIELD_LABELS[field] for field in CREDIT_IMPORT_FIELDS])
    return "\ufeff" + output.getvalue()


def credit_observations_csv(
    items: list[CreditObservationCreate | CreditObservationItem],
) -> str:
    """Export current observations in the same schema accepted by the importer."""
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow([FIELD_LABELS[field] for field in CREDIT_IMPORT_FIELDS])
    status_labels = {
        "active": "正常观察",
        "watch": "重点关注",
        "review": "待复盘",
        "archived": "已归档",
    }
    for item in items:
        payload = item.model_dump(mode="json")
        writer.writerow(
            [
                status_labels.get(str(payload.get(field)), payload.get(field, ""))
                if field == "watch_status"
                else payload.get(field, "")
                for field in CREDIT_IMPORT_FIELDS
            ]
        )
    return "\ufeff" + output.getvalue()


def credit_history_csv(
    items: list[CreditObservationItem], history: list[dict[str, Any]]
) -> str:
    """Export every saved snapshot while carrying current descriptive master fields."""
    current = {item.id: item.model_dump(mode="json") for item in items}
    rows: list[CreditObservationCreate] = []
    ordered_history = sorted(
        history,
        key=lambda snapshot: (
            str(snapshot.get("bond_code") or ""),
            str(snapshot.get("yield_observation_date") or ""),
            str(snapshot.get("recorded_at") or ""),
        ),
    )
    for snapshot in ordered_history:
        master = current.get(int(snapshot["observation_id"]))
        if master is None:
            continue
        payload = {
            # Preserve explicit historical nulls. Only fields that were never part
            # of the history schema (for example sector or watch status) may use
            # the current descriptive master value.
            field: snapshot[field] if field in snapshot else master.get(field)
            for field in CREDIT_IMPORT_FIELDS
        }
        rows.append(CreditObservationCreate(**payload))
    return credit_observations_csv(rows)
